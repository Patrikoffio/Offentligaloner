#!/usr/bin/env python3
"""
Backfill av source_documents.received_at (utlämningsdatum) – hård regel 3.

Prioritet per källdokument:
  1. Kommunlista 2024, kolumn "Datum för mottagen lönelista" (endast riktiga datum).
  2. Datum i filnamnet (file_reference), t.ex. "Österåker 2024-03-16.xlsx".
  3. salary_date (sista fallback).

Kör dry-run som standard; --apply utför UPDATE. Ange --db-url (eller DATABASE_URL).
"""

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path

import openpyxl
import psycopg2

DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
KOMMUNLISTA = Path(__file__).parent.parent / "data" / "stodfiler" / "Kommunlista__2024.xlsx"


def norm(name: str) -> str:
    """Normalisera kommunnamn för matchning: gemener, ta bort 'kommun', trailing s."""
    n = (name or "").lower().strip()
    n = n.replace("kommuns", "").replace("kommun", "").strip()
    n = n.rstrip("s").strip()
    return n


def load_mottagen(path: Path) -> dict[str, dt.date]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dt.date] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in ws.iter_rows(min_row=2):
            kommun = r[0].value if len(r) > 0 else None
            c = r[2].value if len(r) > 2 else None  # "Datum för mottagen lönelista"
            if not kommun:
                continue
            if isinstance(c, dt.datetime):
                out[norm(str(kommun))] = c.date()
            elif isinstance(c, dt.date):
                out[norm(str(kommun))] = c
    return out


def date_from_filename(ref: str) -> dt.date | None:
    m = DATE_RE.search(ref or "")
    if not m:
        return None
    try:
        return dt.date.fromisoformat(m.group(1))
    except ValueError:
        return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db-url", default=os.getenv("DATABASE_URL",
        "postgresql://postgres:postgres@127.0.0.1:54322/postgres"))
    ap.add_argument("--apply", action="store_true", help="Utför UPDATE (annars dry-run)")
    args = ap.parse_args()

    mottagen = load_mottagen(KOMMUNLISTA)
    print(f"Kommunlista: {len(mottagen)} kommuner med giltigt mottaget-datum")

    conn = psycopg2.connect(args.db_url)
    cur = conn.cursor()
    cur.execute("""
        SELECT sd.id, e.name, sd.file_reference, sd.salary_date, sd.received_at
        FROM source_documents sd
        JOIN collection_requests cr ON cr.id = sd.request_id
        JOIN employers e ON e.id = cr.employer_id
        ORDER BY sd.id
    """)
    rows = cur.fetchall()

    counts = {"kommunlista": 0, "filnamn": 0, "salary_date": 0, "redan_satt": 0}
    updates: list[tuple[dt.date, int]] = []
    for sid, ename, ref, sdate, existing in rows:
        if existing is not None:
            counts["redan_satt"] += 1
            continue
        src = mottagen.get(norm(ename))
        if src:
            counts["kommunlista"] += 1
        else:
            src = date_from_filename(ref)
            if src:
                counts["filnamn"] += 1
            else:
                src = sdate
                counts["salary_date"] += 1
        if src:
            updates.append((src, sid))

    print(f"Källor: {counts}")
    print(f"Att uppdatera: {len(updates)} rader")
    if updates[:3]:
        print("Exempel:", updates[:3])

    if args.apply:
        cur.executemany(
            "UPDATE source_documents SET received_at = %s WHERE id = %s AND received_at IS NULL",
            updates,
        )
        conn.commit()
        print(f"APPLICERAT: {cur.rowcount if cur.rowcount != -1 else len(updates)} rader uppdaterade")
    else:
        print("DRY-RUN (ingen ändring). Kör med --apply för att utföra.")
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
